#!/usr/bin/env python3
"""
Consolidate monthly Genesys XLSX files into one XLSX per batch.
Each consolidated file has:
  - Tab 1: "Summary" — combined Apr/May/Jun + Total columns
  - Subsequent tabs: one per month (April, May, June) with individual call paths
"""

import os
from collections import OrderedDict
from datetime import timedelta

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE, "Output")

# Map short month names to display names
MONTH_MAP = {
    "Apr": "April",
    "May": "May",
    "Jun": "June",
    "Jul": "July",
    "Aug": "August",
    "Sep": "September",
    "Oct": "October",
    "Nov": "November",
    "Dec": "December",
    "Jan": "January",
    "Feb": "February",
    "Mar": "March",
}

# Batch file lists in chronological order
BATCHES = {
    "Switchboard": [
        "Switchboard - doc_b696083cdb96_Switchboard-April.xlsx",
        "Switchboard - doc_91b7d70ec942_Switchboard-May.xlsx",
        "Switchboard - doc_73661aaf8493_Switchboard June.xlsx",
    ],
    "Non-ATTY": [
        "Non-ATTY - doc_d98d6ef6429d_Non-ATTY April.xlsx",
        "Non-ATTY - doc_65dfd2d8d7f7_Non-ATTY May.xlsx",
        "Non-ATTY - doc_14b5e25c2c25_Non-ATTY June.xlsx",
    ],
    "ATTY": [
        "ATTY - doc_d0cec11475c8_ATTY April.xlsx",
        "ATTY - doc_38dfa13d8eb9_ATTY May.xlsx",
        "ATTY - doc_74e69e181d95_ATTY June.xlsx",
    ],
    "Bar Examiners": [
        "Bar Examiners - doc_c848269b044c_Bar Examiners April.xlsx",
        "Bar Examiners - doc_7ec1019bfcbe_Bar Examiners May.xlsx",
        "Bar Examiners - doc_a7d02a33e3f6_Bar Examiners June.xlsx",
    ],
    "LRS Spanish": [
        "LRS Spanish - doc_87985000e90b_LRS Spanish April.xlsx",
        "LRS Spanish - doc_66820355f921_LRS Spanish May.xlsx",
        "LRS Spanish - doc_edceaaca07cc_LRS Spanish June.xlsx",
    ],
}

# Metrics order matching the Summary sheet
METRICS = [
    "Total Calls",
    "Successful",
    "Failed",
    "Abandoned",
    "Success %",
    "Abandon %",
    "Average Duration",
]


def extract_month_from_filename(fname):
    """Extract month display name from a filename like 'Switchboard - ... - April.xlsx'."""
    for short, full in MONTH_MAP.items():
        if full in fname:
            return full
        if short in fname and full not in fname:
            return full
    return fname  # fallback


def parse_summary_sheet(filepath):
    """Read the Summary sheet and return a dict of metric -> value."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb["Summary"]
    data = {}
    for row in ws.iter_rows(min_row=2, max_row=10, values_only=False):
        metric = str(row[0].value).strip() if row[0].value else ""
        value = row[1].value
        if metric:
            data[metric] = value
    wb.close()
    return data


def read_call_paths(filepath):
    """Read the Call Paths sheet and return list of (path, count) tuples."""
    wb = load_workbook(filepath, read_only=True)
    ws = wb["Call Paths"]
    paths = []
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    for row in rows:
        path = str(row[0]).strip() if row[0] else ""
        count = row[1]
        if path and count is not None:
            paths.append((path, int(count)))
    wb.close()
    return paths


def format_duration_for_comparison(val):
    """Format duration value for display. Accepts string like '0:06:26' or '00:06:26'."""
    if val is None or val == "N/A":
        return "N/A"
    val = str(val)
    parts = val.split(":")
    if len(parts) == 3:
        h, m, s = parts
        # Normalize: if h has 2 digits and there's no day prefix, strip leading zero
        return val
    return val


def merge_call_paths(all_paths_list):
    """Merge call paths from multiple months. Returns list of (path, apr, may, jun)."""
    path_counts = {}
    for paths in all_paths_list:
        for path, count in paths:
            if path not in path_counts:
                path_counts[path] = {}
            path_counts[path][count] = count  # won't work, need month index
    
    # Actually return list of (path, [counts per month])
    merged = []
    all_unique_paths = []
    seen = set()
    for paths in all_paths_list:
        for path, count in paths:
            if path not in seen:
                all_unique_paths.append(path)
                seen.add(path)
    
    for path in all_unique_paths:
        counts = []
        for paths in all_paths_list:
            path_dict = dict(paths)
            counts.append(path_dict.get(path, 0))
        merged.append((path, counts))
    
    return merged


def write_consolidated(batch_name, monthly_files, output_path):
    """Write the consolidated XLSX for a batch."""
    # Read all monthly data
    monthly_data = []  # list of (month_name, metrics_dict, paths_list)
    for fname in monthly_files:
        fpath = os.path.join(OUT_DIR, fname)
        if not os.path.exists(fpath):
            print(f"  ⚠ Missing: {fname}")
            continue
        month = extract_month_from_filename(fname)
        metrics = parse_summary_sheet(fpath)
        paths = read_call_paths(fpath)
        monthly_data.append((month, metrics, paths))
    
    if not monthly_data:
        print(f"  ⚠ No valid files found for {batch_name}")
        return

    # Create a fresh workbook
    wb = Workbook()
    # Rename the default sheet (Sheet) to "Summary"
    default_name = wb.sheetnames[0] if wb.sheetnames else "Sheet"
    ws = wb[default_name]
    ws.title = "Summary"
    
    # Columns: Metric, April, May, June, Total
    months_display = [m[0] for m in monthly_data]  # ['April', 'May', 'June']
    headers = ["Metric"] + months_display + ["Total"]
    ws["A1"] = headers[0]
    for col, h in enumerate(headers[1:], start=2):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
    
    # Data rows — metrics as rows, months + Total as columns
    # For numeric metrics, sum across months for Total; for percentages, average; for strings, pick latest
    month_indices = {m: i for i, m in enumerate(months_display)}
    
    for row_idx, metric_name in enumerate(METRICS, start=2):
        ws.cell(row=row_idx, column=1, value=metric_name)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        
        values_across_months = []
        for col_idx, (month, metrics, _) in enumerate(monthly_data, start=2):
            val = metrics.get(metric_name)
            ws.cell(row=row_idx, column=col_idx, value=val)
            
            if metric_name in ("Total Calls", "Successful", "Failed", "Abandoned"):
                values_across_months.append(int(val))
            elif metric_name in ("Success %", "Abandon %"):
                # Strip % sign, convert to float
                if isinstance(val, str) and val.endswith('%'):
                    values_across_months.append(float(val.replace('%', '')))
                else:
                    values_across_months.append(float(val))
            elif metric_name == "Average Duration":
                # Parse duration string to seconds for averaging
                if val and val != "N/A":
                    val_str = str(val)
                    parts = val_str.split(":")
                    if len(parts) == 3:
                        h, m, s = parts
                        secs = int(h) * 3600 + int(m) * 60 + float(s)
                        values_across_months.append(secs)
                    else:
                        values_across_months.append(0)
                else:
                    values_across_months.append(0)
            else:
                # String values (Call Path, Month) — just keep the latest month's value
                values_across_months.append(val)
        
        # Calculate Total column
        if metric_name in ("Total Calls", "Successful", "Failed", "Abandoned"):
            ws.cell(row=row_idx, column=len(headers), value=sum(values_across_months))
        elif metric_name in ("Success %", "Abandon %"):
            avg = sum(values_across_months) / len(values_across_months) if values_across_months else 0
            ws.cell(row=row_idx, column=len(headers), value=f"{avg:.2f}%")
        elif metric_name == "Average Duration":
            avg_secs = sum(values_across_months) / len(values_across_months) if values_across_months else 0
            td = str(timedelta(seconds=round(avg_secs)))
            ws.cell(row=row_idx, column=len(headers), value=td)
        else:
            ws.cell(row=row_idx, column=len(headers), value=values_across_months[-1] if values_across_months else "")
    
    # Column widths
    ws.column_dimensions["A"].width = 22
    for i in range(len(monthly_data) + 2):
        col_letter = chr(66 + i)
        ws.column_dimensions[col_letter].width = 16
    
    # --- Tabs for each month's Call Paths ---
    for month_idx, (month, metrics, paths) in enumerate(monthly_data):
        ws2 = wb.create_sheet(title=month)
        ws2["A1"] = "Path"
        ws2["B1"] = "Count"
        ws2["A1"].font = Font(bold=True)
        ws2["B1"].font = Font(bold=True)
        ws2["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws2["A1"].font = Font(bold=True, color="FFFFFF")
        ws2["B1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws2["B1"].font = Font(bold=True, color="FFFFFF")
        
        for row_idx, (path, count) in enumerate(paths, start=2):
            ws2.cell(row=row_idx, column=1, value=path)
            ws2.cell(row=row_idx, column=2, value=count)
        
        ws2.column_dimensions["A"].width = 80
        ws2.column_dimensions["B"].width = 10
    
    wb.save(output_path)
    print(f"  ✓ {output_path}")


def main():
    for batch_name, files in BATCHES.items():
        print(f"\nConsolidating: {batch_name}")
        out_name = f"{batch_name} - Consolidated.xlsx"
        out_path = os.path.join(OUT_DIR, out_name)
        write_consolidated(batch_name, files, out_path)
    
    print(f"\nDone! All consolidated files in: {OUT_DIR}")


if __name__ == "__main__":
    main()
