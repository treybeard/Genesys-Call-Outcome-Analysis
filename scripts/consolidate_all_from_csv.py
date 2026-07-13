#!/usr/bin/env python3
"""
consolidate_all_from_csv.py

Reads ALL CSV files from the csv files directory, groups them by category,
and writes one consolidated Excel per category with:
  - Summary tab  (Total Calls)
  - Monthly tabs (April, May, June) with call paths + counts + bar charts

Uses openpyxl for tables and bar charts.
Based on scripts/analyse_calls.py and scripts/consolidate_call_flow_reports.py.
"""

import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_DIR = os.path.join(PROJECT_ROOT, "CSV Files")
OUT_DIR = os.path.join(PROJECT_ROOT, "Output")

MONTH_NAMES = ["April", "May", "June"]


def discover_categories():
    """
    Scan CSV_DIR and group files by category (prefix before month name).
    Returns: { category_name: [(month, filename), ...] }
    """
    csv_files = [f for f in os.listdir(CSV_DIR) if f.lower().endswith(".csv")]
    groups = {}

    for fname in csv_files:
        match = re.search(r"(april|may|june)", fname, re.IGNORECASE)
        if not match:
            print(f"  ⚠ No month found in '{fname}', skipping")
            continue

        month = match.group(1).capitalize()
        # Category = everything before the month name, stripping doc_ prefix
        prefix = fname[:match.start()].strip()
        category = re.sub(r"^doc_[a-f0-9]+_", "", prefix).strip()
        # Normalize typos: "Switchbaord" or "Switchbaord-" → "Switchboard"
        category = re.sub(r'(?i)switchboa?r[dh]|switchba?o?r?d', 'Switchboard', category).strip()
        category = category.strip('-').strip()  # strip trailing hyphens/spaces

        if category not in groups:
            groups[category] = {}
        groups[category][month] = fname

    # Convert to ordered list (April, May, June)
    result = {}
    for cat, months_dict in groups.items():
        ordered = []
        for m in MONTH_NAMES:
            if m in months_dict:
                ordered.append((m, months_dict[m]))
        if len(ordered) >= 1:
            result[cat] = ordered

    return result


def count_queues_for_csv(csv_path):
    """Read one CSV, return (total_calls, queue_df sorted by count desc)."""
    df = pd.read_csv(csv_path, dtype=str)
    df = df.dropna(axis=1, how='all')
    total_calls = len(df)

    queue_col = next((col for col in df.columns if col.lower() == 'queue'), None)
    if queue_col is None:
        return total_calls, pd.DataFrame(columns=['Queue Path', 'Call Count'])

    queue_counts = df[queue_col].value_counts(dropna=True)
    queue_df = queue_counts.reset_index()
    queue_df.columns = ['Queue Path', 'Call Count']
    queue_df = queue_df.sort_values('Call Count', ascending=False).reset_index(drop=True)
    return total_calls, queue_df


# ─── Styles (from scripts/consolidate_call_flow_reports.py) ─────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def add_monthly_chart(ws_m, sheet_name):
    """Add bar chart below monthly sheet's data."""
    chart = BarChart()
    chart.title = f"{sheet_name} Call Paths"
    chart.y_axis.title = "Call Count"
    chart.width = 15
    chart.height = 12
    chart.style = 1  # 2D look (no 3D)

    labels = Reference(ws_m, min_col=1, min_row=2, max_row=ws_m.max_row)
    data = Reference(ws_m, min_col=2, min_row=2, max_row=ws_m.max_row)

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    anchor_row = ws_m.max_row + 5
    ws_m.add_chart(chart, f"A{anchor_row}")
    return anchor_row


def process_category(cat_name, csv_entries):
    """Process one category and write consolidated Excel."""
    print(f"\n{'='*60}")
    print(f"Processing: {cat_name}")
    print(f"{'='*60}")

    monthly_data = {}  # month → (total, queue_df)
    total_all = 0

    for month_name, fname in csv_entries:
        csv_full = os.path.join(CSV_DIR, fname)
        if not os.path.exists(csv_full):
            print(f"  ⚠ Missing: {fname}")
            continue

        total, qdf = count_queues_for_csv(csv_full)
        monthly_data[month_name] = (total, qdf)
        total_all += total
        print(f"  {month_name}: {total} calls, {len(qdf)} unique paths")

    if not monthly_data:
        print("  ❌ No data found, skipping.")
        return False

    out_path = os.path.join(OUT_DIR, f"{cat_name} - Consolidated.xlsx")
    os.makedirs(OUT_DIR, exist_ok=True)
    print(f"  Writing: {out_path}")

    # Build workbook (openpyxl doesn't have "create new" from thin air, so we
    # use openpyxl.Workbook() directly)
    from openpyxl import Workbook
    wb = Workbook()

    # ── Summary sheet ──
    headers = ['Metric'] + list(monthly_data.keys()) + ['Total']
    ws_sum = wb.create_sheet('Summary')

    for ci, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws_sum.column_dimensions['A'].width = 24
    for ci in range(2, len(headers) + 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 14

    # Row 2: Total Calls
    row = 2
    ws_sum.cell(row=row, column=1, value='Total Calls').border = THIN_BORDER
    for ci, m_name in enumerate(monthly_data.keys(), start=2):
        ws_sum.cell(row=row, column=ci, value=monthly_data[m_name][0]).border = THIN_BORDER
    ws_sum.cell(row=row, column=len(headers), value=total_all).border = THIN_BORDER
    ws_sum.cell(row=row, column=len(headers)).fill = TOTAL_FILL

    wb.save(out_path)

    # ── Reopen and add monthly sheets with charts ──
    wb = load_workbook(out_path)

    for month_name, (total, qdf) in monthly_data.items():
        sheet_name = month_name

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws_m = wb.create_sheet(sheet_name)

        # Header
        ws_m.cell(row=1, column=1, value='Queue Path').font = HEADER_FONT
        ws_m.cell(row=1, column=1).fill = HEADER_FILL
        ws_m.cell(row=1, column=1).alignment = HEADER_ALIGN
        ws_m.cell(row=1, column=1).border = THIN_BORDER
        ws_m.cell(row=1, column=2, value='Call Count').font = HEADER_FONT
        ws_m.cell(row=1, column=2).fill = HEADER_FILL
        ws_m.cell(row=1, column=2).alignment = HEADER_ALIGN
        ws_m.cell(row=1, column=2).border = THIN_BORDER

        # Data rows
        for pi, (_, row_data) in enumerate(qdf.iterrows(), start=2):
            ws_m.cell(row=pi, column=1, value=row_data['Queue Path']).border = THIN_BORDER
            ws_m.cell(row=pi, column=2, value=int(row_data['Call Count'])).border = THIN_BORDER

        # Format
        for r in range(1, ws_m.max_row + 1):
            for c in range(1, ws_m.max_column + 1):
                cell = ws_m.cell(row=r, column=c)
                cell.border = THIN_BORDER
                if r == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = HEADER_ALIGN

        ws_m.column_dimensions['A'].width = 55
        if ws_m.max_column >= 2:
            ws_m.column_dimensions[get_column_letter(2)].width = 12

        # Add bar chart
        anchor = add_monthly_chart(ws_m, sheet_name)
        print(f"  ✓ {sheet_name}: {len(qdf)} paths + chart at row {anchor}")

    wb.save(out_path)
    print(f"  ✓ Final saved: {out_path}")
    return True


def main():
    print("=" * 60)
    print("Genesys Call Flow: CSV → Consolidated Excel (ALL CATEGORIES)")
    print(f"CSV directory: {CSV_DIR}")
    print(f"Output directory: {OUT_DIR}")
    print("=" * 60)

    categories = discover_categories()
    print(f"\nDiscovered {len(categories)} categories: {', '.join(categories.keys())}")

    results = {}
    for cat_name, csv_entries in categories.items():
        try:
            success = process_category(cat_name, csv_entries)
            results[cat_name] = '✓' if success else '✗'
        except Exception as e:
            print(f"  ❌ Error processing {cat_name}: {e}")
            import traceback
            traceback.print_exc()
            results[cat_name] = '✗'

    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    for cat, status in results.items():
        print(f"  {status} {cat}")

    success_count = sum(1 for s in results.values() if s == '✓')
    print(f"\n✅ {success_count}/{len(results)} categories processed")


if __name__ == '__main__':
    main()
