#!/usr/bin/env python3
"""
consolidate_switchboard_from_csv.py

Takes the 3 monthly Switchboard CSV exports, counts unique call paths per month,
and writes ONE consolidated Excel with:
  - Summary tab  (total calls across all months)
  - April tab  (call paths + counts for April)
  - May tab    (call paths + counts for May)
  - June tab   (call paths + counts for June)

Uses the logic from analyse_calls.py (found in scripts/).
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_DIR = os.path.join(PROJECT_ROOT, "CSV Files")
OUT_DIR = os.path.join(PROJECT_ROOT, "Output")

# The 3 Switchboard CSV files (ordered)
CSV_FILES = [
    ("April",  "doc_b696083cdb96_Switchbaord-April.csv"),
    ("May",    "doc_91b7d70ec942_Switchbaord-May.csv"),
    ("June",   "doc_73661aaf8493_Switchboard June.csv"),
]

# ─── Styles ─────────────────────────────────────────────────────────────

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
TOTAL_FILL   = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FONT   = Font(bold=True)
THIN_BORDER  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)


# ─── Core logic (from analyse_calls.py) ─────────────────────────────────

def count_queues_for_csv(csv_path):
    """Read one CSV, return (total_calls, queue_path_df)."""
    df = pd.read_csv(csv_path, dtype=str)
    df = df.dropna(axis=1, how='all')
    total_calls = len(df)

    queue_col = next((col for col in df.columns if col.lower() == 'queue'), None)
    if queue_col is None:
        # No queue column → nothing to report
        return total_calls, pd.DataFrame(columns=['Queue Path', 'Call Count'])

    queue_counts = df[queue_col].value_counts(dropna=True)
    queue_df = queue_counts.reset_index()
    queue_df.columns = ['Queue Path', 'Call Count']
    queue_df = queue_df.sort_values('Call Count', ascending=False).reset_index(drop=True)
    return total_calls, queue_df


# ─── Build consolidated Excel ───────────────────────────────────────────

def build_switchboard_xlsx():
    csv_path_all = []
    all_months = {}  # month_name → (total, queue_df)

    for month_name, fname in CSV_FILES:
        csv_full = os.path.join(CSV_DIR, fname)
        csv_path_all.append(csv_full)
        total, qdf = count_queues_for_csv(csv_full)
        all_months[month_name] = (total, qdf)
        print(f"  {month_name}: {total} calls, {len(qdf)} unique paths")

    total_all = sum(m[0] for m in all_months.values())

    # ── Write workbook with openpyxl ───────────────────────────────────
    out_path = os.path.join(OUT_DIR, "Switchboard - Consolidated.xlsx")
    os.makedirs(OUT_DIR, exist_ok=True)

    wb = load_workbook()  # start fresh — we'll build from scratch

    # Summary sheet
    ws_sum = wb.create_sheet('Summary')
    headers = ['Metric'] + list(all_months.keys()) + ['Total']
    for ci, h in enumerate(headers, 1):
        cell = ws_sum.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws_sum.column_dimensions['A'].width = 22
    for ci in range(2, len(headers) + 1):
        col_letter = get_column_letter(ci)
        ws_sum.column_dimensions[col_letter].width = 14

    # Total column values
    total_calls = sum(m[0] for m in all_months.values())

    # Row 2: Total Calls
    row = 2
    ws_sum.cell(row=row, column=1, value='Total Calls').border = THIN_BORDER
    for ci, m_name in enumerate(all_months.keys(), start=2):
        ws_sum.cell(row=row, column=ci, value=all_months[m_name][0]).border = THIN_BORDER
    ws_sum.cell(row=row, column=len(headers), value=total_calls).border = THIN_BORDER
    ws_sum.cell(row=row, column=len(headers)).fill = TOTAL_FILL
    ws_sum.cell(row=row, column=len(headers)).font = TOTAL_FONT
    row += 1

    # Accumulate all queues for Summary-level unique queue list (just all unique paths)
    all_queue_paths = set()
    for m_name, (total, qdf) in all_months.items():
        all_queue_paths.update(qdf['Queue Path'].tolist())

    print(f"\n  Writing: {out_path}")
    wb.save(out_path)
    print(f"  ✓ Saved (Summary + monthly sheets)")

    # Now reopen and add charts + format monthly sheets
    wb = load_workbook(out_path)

    # Add Summary totals row (Grand Total)
    ws = wb['Summary']
    # Already done above, now add "Unique Queue Paths" metric
    unique_count = len(all_queue_paths)
    row += 1
    ws.cell(row=row, column=1, value='Unique Queue Paths').border = THIN_BORDER
    for ci, m_name in enumerate(all_months.keys(), start=2):
        ws.cell(row=row, column=ci, value=len(all_months[m_name][1])).border = THIN_BORDER
    ws.cell(row=row, column=len(headers), value=unique_count).border = THIN_BORDER
    ws.cell(row=row, column=len(headers)).fill = TOTAL_FILL

    # Format monthly sheets + add charts
    for month_name in all_months:
        sheet_name = month_name  # "April", "May", "June"
        if sheet_name not in wb.sheetnames:
            continue

        ws_m = wb[sheet_name]

        # Format: borders + header
        for r in range(1, ws_m.max_row + 1):
            for c in range(1, ws_m.max_column + 1):
                cell = ws_m.cell(row=r, column=c)
                cell.border = THIN_BORDER
                if r == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = HEADER_ALIGN

        ws_m.column_dimensions['A'].width = 55  # call paths can be long
        if ws_m.max_column >= 2:
            ws_m.column_dimensions[get_column_letter(2)].width = 12

        # Add bar chart
        chart = BarChart()
        chart.title = f"{sheet_name} Call Paths"
        chart.y_axis.title = "Call Count"
        chart.style = 10
        chart.width = 15  # ~15cm
        chart.height = 12  # ~12cm

        # Data starts from row 2 (skip header)
        labels = Reference(ws_m, min_col=1, min_row=2, max_row=ws_m.max_row)
        data = Reference(ws_m, min_col=2, min_row=2, max_row=ws_m.max_row)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)

        anchor_row = ws_m.max_row + 5
        ws_m.add_chart(chart, f"A{anchor_row}")

        print(f"  ✓ {sheet_name}: {len(ws_m) - 1} paths + chart at row {anchor_row}")

    wb.save(out_path)
    print(f"  ✓ Final saved: {out_path}")


if __name__ == '__main__':
    build_switchboard_xlsx()
