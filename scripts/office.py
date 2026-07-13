#!/usr/bin/env python3
"""
office.py — Excel workbook generation for Genesys call analysis.

Two modes:
  1. simple_xlsx() — Data-only workbooks (no charts) via openpyxl.
  2. full_xlsx() — Full workbooks with Summary + monthly tabs + bar charts
     (hand-crafted XML chart/drawing, no openpyxl charts).

CRITICAL: Do NOT use openpyxl.chart.BarChart — it produces XML that
Excel silently hides. Always hand-craft chart XML + drawing XML.
"""

import csv
import os
import zipfile
from datetime import timedelta
from collections import Counter
from engineering import (
    BATCHES, analyze_csv, analyze_batch, weight_avg,
    LRS_SPANISH_BATES,
)


# ── Constants (matching skill spec) ────────────────────────────────────────
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
DRAWING_NS = (
    "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
)
RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
THEME_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"

CHART_WIDTH_EMU = 5_400_000_000   # ~54mm
ROW_HEIGHT_EMU = 68_580_000       # per data row
OFFSET_EMU = 12_700               # top/left offset
TO_OFFSET_EMU = 177_800           # bottom offset


# ── Helpers ────────────────────────────────────────────────────────────────

def _ns(tag, ns=CHART_NS):
    return f"ns0:{tag}"


def _xml_escape(s):
    return (str(s).replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&apos;"))


def _a_tag(tag):
    """Wrap a tag in a: namespace."""
    return f"a:{tag}"


def _xdr_tag(tag):
    """Wrap a tag in xdr: namespace."""
    return f"xdr:{tag}"


def _cell_ref(row, col):
    """1-based row, col → Excel A1 notation."""
    result = ""
    c = col
    while c > 0:
        c, remainder = divmod(c - 1, 26)
        result = chr(65 + remainder) + result
    return result + str(row)


# ── Simple openpyxl workbooks (no charts) ─────────────────────────────────

def simple_xlsx(batch_name, monthly_data, output_path):
    """Write a data-only consolidated XLSX (no charts).

    Args:
        batch_name: e.g. 'Switchboard'
        monthly_data: {month: analysis_dict}
        output_path: where to write the .xlsx file.

    Writes: Summary sheet (metric × months + Total column) + monthly tabs.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl is required for simple_xlsx")

    months_order = list(monthly_data.keys())
    # Enforce chronological order (Apr, May, Jun, ...)
    MONTH_ORDER = ["April", "May", "June", "July", "August",
                   "September", "October", "November", "December",
                   "January", "February", "March"]
    months_order.sort(key=lambda m: MONTH_ORDER.index(m)
                      if m in MONTH_ORDER else 999)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Header
    headers = ["Metric"] + months_order + ["Total"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = __import__("openpyxl").styles.Font(
            bold=True, color="FFFFFF", size=11
        )
        cell.fill = __import__("openpyxl").styles.PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        cell.alignment = __import__("openpyxl").styles.Alignment(
            horizontal="center", vertical="center"
        )
        cell.border = __import__("openpyxl").styles.Border(
            left=__import__("openpyxl").styles.Side(style="thin"),
            right=__import__("openpyxl").styles.Side(style="thin"),
            top=__import__("openpyxl").styles.Side(style="thin"),
            bottom=__import__("openpyxl").styles.Side(style="thin"),
        )
    ws.column_dimensions["A"].width = 22
    for ci in range(2, len(headers) + 1):
        col_letter = __import__("openpyxl").utils.get_column_letter(ci)
        ws.column_dimensions[col_letter].width = 16

    # Metrics rows — use snake_case keys that match engineering.py output
    METRICS = [
        "Total Calls", "Successful", "Failed", "Abandoned",
        "Success %", "Abandon %", "Average Duration",
    ]
    # Mapping from display name → engine dict key
    METRIC_KEYS = {
        "Total Calls": "total_calls",
        "Successful": "successful",
        "Failed": "failed",
        "Abandoned": "abandoned",
        "Success %": "success_pct",
        "Abandon %": "abandon_pct",
        "Average Duration": "avg_duration",
    }
    # Display values for Total label row
    DISPLAY_NAMES = {
        "total_calls": "Total Calls",
        "successful": "Successful",
        "failed": "Failed",
        "abandoned": "Abandoned",
    }

    total_col = len(months_order) + 2  # Total column index (1-based)

    for ri, metric_name in enumerate(METRICS, start=2):
        ws.cell(row=ri, column=1, value=metric_name).font = (
            __import__("openpyxl").styles.Font(bold=True)
        )
        values_across_months = []
        for ci, m_name in enumerate(months_order, start=2):
            d = monthly_data[m_name]
            engine_key = METRIC_KEYS[metric_name]
            val = d.get(engine_key)
            cell = ws.cell(row=ri, column=ci)

            # Accumulate for Total column
            if metric_name in (
                "Total Calls", "Successful", "Failed", "Abandoned"
            ):
                cell.value = int(val or 0)
                values_across_months.append(int(val or 0))
            elif metric_name in ("Success %", "Abandon %"):
                # engineering.py returns 0-100 range; write as decimal (0-1)
                pct_val = (val or 0) / 100
                cell.value = pct_val
                cell.number_format = "0.00%"
                values_across_months.append(pct_val)
            elif metric_name == "Average Duration":
                if val and val != "N/A" and val != "0:00:00":
                    # Could be float (seconds) or str ("mm:ss")
                    if isinstance(val, (int, float)):
                        values_across_months.append(val)
                        cell.value = val
                    else:
                        td_parts = str(val).split(":")
                        if len(td_parts) == 3:
                            secs = (
                                int(td_parts[0]) * 3600
                                + int(td_parts[1]) * 60
                                + float(td_parts[2])
                            )
                            values_across_months.append(secs)
                            cell.value = secs
                        else:
                            values_across_months.append(0)
                            cell.value = 0
                else:
                    values_across_months.append(0)
                    cell.value = 0
            else:
                cell.value = val
                values_across_months.append(val)

        # Total column calculation
        total_r = _cell_ref(ri, total_col)
        if metric_name in (
            "Total Calls", "Successful", "Failed", "Abandoned"
        ):
            ws.cell(row=ri, column=total_col, value=sum(values_across_months))
        elif metric_name in ("Success %", "Abandon %"):
            avg = (
                sum(values_across_months) / len(values_across_months)
                if values_across_months else 0
            )
            # Write as decimal 0-1 with percentage format (NOT a string)
            total_cell = ws.cell(row=ri, column=total_col)
            total_cell.value = avg
            total_cell.number_format = "0.00%"
        elif metric_name == "Average Duration":
            avg_secs = (
                sum(values_across_months) / len(values_across_months)
                if values_across_months else 0
            )
            ws.cell(row=ri, column=total_col, value=str(timedelta(
                seconds=round(avg_secs)
            )))
        else:
            ws.cell(row=ri, column=total_col, value=values_across_months[-1]
                    if values_across_months else "")

    # Monthly tabs
    for month_idx, m_name in enumerate(months_order):
        d = monthly_data[m_name]
        ws2 = wb.create_sheet(title=m_name)
        ws2["A1"] = "Path"
        ws2["B1"] = "Count"

        header_font = __import__("openpyxl").styles.Font(
            bold=True, color="FFFFFF", size=11
        )
        header_fill = __import__("openpyxl").styles.PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        thin_border = __import__("openpyxl").styles.Border(
            left=__import__("openpyxl").styles.Side(style="thin"),
            right=__import__("openpyxl").styles.Side(style="thin"),
            top=__import__("openpyxl").styles.Side(style="thin"),
            bottom=__import__("openpyxl").styles.Side(style="thin"),
        )

        ws2["A1"].font = header_font
        ws2["A1"].fill = header_fill
        ws2["A1"].alignment = __import__("openpyxl").styles.Alignment(
            horizontal="center", vertical="center"
        )
        ws2["A1"].border = thin_border

        ws2["B1"].font = header_font
        ws2["B1"].fill = header_fill
        ws2["B1"].alignment = __import__("openpyxl").styles.Alignment(
            horizontal="center", vertical="center"
        )
        ws2["B1"].border = thin_border

        for pi, (path, count) in enumerate(d["call_paths"], start=2):
            ws2.cell(row=pi, column=1, value=path)
            ws2.cell(row=pi, column=2, value=count)
            ws2.cell(row=pi, column=1).border = thin_border
            ws2.cell(row=pi, column=2).border = thin_border

        ws2.column_dimensions["A"].width = 80
        ws2.column_dimensions["B"].width = 10

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# ── Full XLSX with manual XML charts ──────────────────────────────────────

def full_xlsx(batch_name, monthly_data, output_path):
    """Build a complete XLSX with Summary + monthly tabs + bar charts.

    Charts are hand-crafted via raw XML (no openpyxl.chart.BarChart).

    Args:
        batch_name: e.g. 'Switchboard'
        monthly_data: {month: analysis_dict}
        output_path: where to write the .xlsx file.
    """
    months_order = list(monthly_data.keys())
    MONTH_ORDER = [
        "April", "May", "June", "July", "August", "September",
        "October", "November", "December", "January", "February", "March"
    ]
    months_order.sort(key=lambda m: MONTH_ORDER.index(m)
                      if m in MONTH_ORDER else 999)

    if not months_order:
        print(f"  ⚠ No valid data for {batch_name}")
        return

    num_months = len(months_order)
    num_sheets = 1 + num_months  # Summary + months
    total_col = num_months + 2  # Total column index (1-based)

    # Collect all unique strings for sharedStrings.xml
    all_texts = set()
    all_texts.update(["Metric"] + months_order + ["Total"])
    for m in months_order:
        all_texts.update(
            ["Total Calls", "Successful", "Failed", "Abandoned",
             "Success %", "Abandon %", "Average Duration"]
        )
    all_texts.update(["Path", "Count"])
    for m in months_order:
        all_texts.update(p for p, _ in monthly_data[m]["call_paths"])

    texts_list = sorted(all_texts)
    shared_xml, str_to_idx = _build_shared_strings(texts_list)

    # ── Summary sheet XML ───────────────────────────────────────────────
    sheet1_data = _build_summary_xml(months_order, monthly_data,
                                     total_col, str_to_idx)

    # ── Monthly sheet XML ───────────────────────────────────────────────
    chart_info = []  # (month_name, paths, counts, num_paths)
    for m_name in months_order:
        d = monthly_data[m_name]
        chart_info.append(
            (m_name,
             [p for p, _ in d["call_paths"]],
             [c for _, c in d["call_paths"]],
             len(d["call_paths"]))
        )

    monthly_sheets = []  # (filename, xml_bytes)
    for sheet_idx, (m_name, paths, counts, num_paths) in enumerate(
        chart_info, start=2
    ):
        sheet_xml = _build_monthly_sheet(m_name, paths, counts, num_paths,
                                         str_to_idx=str_to_idx)
        monthly_sheets.append(
            (f"sheet{sheet_idx}.xml", sheet_xml)
        )

    # ── Chart XML ───────────────────────────────────────────────────────
    chart_files = []  # (chart_filename, xml_bytes, month_name)
    for i, (m_name, paths, counts, num_paths) in enumerate(
        chart_info, start=1
    ):
        chart_xml = _build_chart_xml(
            i, m_name, paths, counts, num_paths
        )
        chart_files.append((f"chart{i}.xml", chart_xml, m_name))

    # ── Drawing XML (both anchors per spec) ─────────────────────────────
    drawing_files = []  # (drawing_filename, xml_bytes, month_name)
    for i, (m_name, paths, counts, num_paths) in enumerate(
        chart_info, start=1
    ):
        drawing_xml = _build_drawing_xml(
            i, m_name, num_paths
        )
        drawing_files.append((f"drawing{i}.xml", drawing_xml, m_name))

    sheet_rels = [(i + 1, f"drawing{i}.xml")
                  for i in range(num_months)]  # (sheet_num, drawing_fname)

    # ── Supporting files ────────────────────────────────────────────────
    content_types = _build_content_types(num_sheets, chart_files,
                                         drawing_files, sheet_rels)
    workbook_xml = _build_workbook(months_order)
    workbook_rels = _build_workbook_rels(months_order)
    styles_xml = _build_styles_xml()
    theme_xml = _build_theme_xml()
    root_rels = _build_root_rels()
    docprops_core = _build_docprops_core()
    docprops_app = _build_docprops_app(months_order)

    # ── Package as xlsx (ZIP) ───────────────────────────────────────────
    out_dir = os.path.dirname(output_path)
    tmp = os.path.join(out_dir, f".tmp_{os.path.basename(output_path)}")
    os.makedirs(out_dir, exist_ok=True)

    with zipfile.ZipFile(tmp, "w") as zf:
        def w(name, content):
            if isinstance(content, bytes):
                zf.writestr(name, content)
            else:
                zf.writestr(name, content.encode("utf-8"))

        # All *_build_* helpers already return bytes; w() handles both
        w("[Content_Types].xml", content_types)
        w("_rels/.rels", root_rels)
        w("docProps/app.xml", docprops_app)
        w("docProps/core.xml", docprops_core)
        w("xl/workbook.xml", workbook_xml)
        w("xl/_rels/workbook.xml.rels", workbook_rels)
        w("xl/styles.xml", styles_xml)
        w("xl/theme/theme1.xml", theme_xml)
        w("xl/sharedStrings.xml", shared_xml)
        w("xl/worksheets/sheet1.xml", sheet1_data)
        for fname, content in monthly_sheets:
            w(f"xl/worksheets/{fname}", content)
        for fname, content, _ in chart_files:
            w(f"xl/charts/{fname}", content)
        for fname, content, _ in drawing_files:
            w(f"xl/drawings/{fname}", content)

        # Chart rels (already bytes)
        for i, (fname, content, _) in enumerate(chart_files, start=1):
            chart_rels = _build_chart_rels(i)
            w(f"xl/charts/_rels/{fname.replace('.xml', '')}.xml.rels",
              chart_rels)

        # Drawing rels (already bytes)
        for i, (fname, content, _) in enumerate(drawing_files, start=1):
            drawing_rels = _build_drawing_rels(i)
            w(f"xl/drawings/_rels/{fname.replace('.xml', '')}.xml.rels",
              drawing_rels)

        # Sheet rels (inline string → encode manually)
        for sheet_num, drawing_fname in sheet_rels:
            sheet_rls_content = (
                f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                f'<Relationships xmlns="{RELS_NS}">'
                f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                f'Target="../drawings/{drawing_fname}"/>'
                f'</Relationships>'
            )
            w(f"xl/worksheets/_rels/sheet{sheet_num}.xml.rels",
              sheet_rls_content.encode("utf-8"))

    os.rename(tmp, output_path)
    print(f"  ✓ {batch_name}: {output_path}")


# ── XML Builders ───────────────────────────────────────────────────────────

def _build_shared_strings(texts):
    """Build sharedStrings.xml. Returns (xml_bytes, {text: index})."""
    seen = {}
    si_list = []
    for t in texts:
        if t not in seen:
            seen[t] = len(si_list)
            si_list.append(t)

    inner = "".join(
        f'<si><t>{_xml_escape(s)}</t></si>' for s in si_list
    )
    xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<sst xmlns="{MAIN_NS}" count="{len(si_list)}" '
        f'uniqueCount="{len(si_list)}">{inner}</sst>'
    )
    return xml.encode("utf-8"), seen


def _build_summary_xml(months_order, monthly_data, total_col, str_to_idx):
    """Build Summary sheet XML."""
    num_rows = len(months_order) + 1  # header + 7 metrics

    parts = []
    parts.append(
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<worksheet xmlns="{MAIN_NS}" '
        f'xmlns:r="{RELS_NS}">'
    )
    parts.append(
        f'<dimension ref="A1:{_cell_ref(num_rows, total_col)}"/>'
    )
    parts.append('<sheetViews><sheetView workbookViewId="0"/></sheetViews>')
    parts.append('<sheetFormatPr baseColWidth="10" defaultRowHeight="15"/>')

    parts.append('<cols><col min="1" max="1" width="22" customWidth="1"/>')
    for i in range(2, total_col + 1):
        parts.append(f'<col min="{i}" max="{i}" width="16" customWidth="1"/>')
    parts.append("</cols>")

    parts.append("<sheetData>")

    # Header row
    header_row = (
        f'<row r="1" spans="1:{total_col}">'
        f'<c r="A1" s="1" t="s"><v>{str_to_idx["Metric"]}</v></c>'
    )
    for ci, m_name in enumerate(months_order, start=2):
        header_row += (
            f'<c r="{_cell_ref(1, ci)}" s="1" t="s">'
            f'<v>{str_to_idx[m_name]}</v></c>'
        )
    header_row += (
        f'<c r="{_cell_ref(1, total_col)}" s="1" t="s">'
        f'<v>{str_to_idx["Total"]}</v></c></row>'
    )
    parts.append(header_row)

    # Metric rows
    METRICS = [
        "Total Calls", "Successful", "Failed", "Abandoned",
        "Success %", "Abandon %", "Average Duration",
    ]
    # Mapping from display name → engine dict key (snake_case)
    METRIC_KEYS = {
        "Total Calls": "total_calls",
        "Successful": "successful",
        "Failed": "failed",
        "Abandoned": "abandoned",
        "Success %": "success_pct",
        "Abandon %": "abandon_pct",
        "Average Duration": "avg_duration",
    }

    for ri, metric_name in enumerate(METRICS, start=2):
        engine_key = METRIC_KEYS[metric_name]
        row_cells = f'<row r="{ri}" spans="1:{total_col}">'
        row_cells += (
            f'<c r="A{ri}" s="2" t="s">'
            f'<v>{str_to_idx[metric_name]}</v></c>'
        )

        values_across_months = []
        for ci, m_name in enumerate(months_order, start=2):
            d = monthly_data[m_name]
            val = d.get(engine_key)
            r = _cell_ref(ri, ci)
            if metric_name in ("Success %", "Abandon %"):
                # engineering.py returns 0-100 range; numFmtId="10" is Excel
                # percentage (multiplies stored value by 100), so store 0-1
                pct_val = float(val or 0) / 100
                row_cells += (
                    f'<c r="{r}" s="4" t="n">'
                    f'<v>{pct_val:.10f}</v></c>'
                )
            elif metric_name == "Average Duration":
                avg_secs = d.get("avg_duration", 0)
                rounded_secs = round(avg_secs)
                row_cells += (
                    f'<c r="{r}" s="5" t="n">'
                    f'<v>{rounded_secs / 86400:.15f}</v></c>'
                )
            else:
                row_cells += (
                    f'<c r="{r}" s="3" t="n">'
                    f'<v>{val}</v></c>'
                )

            # Accumulate for Total
            if metric_name in (
                "Total Calls", "Successful", "Failed", "Abandoned"
            ):
                values_across_months.append(int(val or 0))
            elif metric_name in ("Success %", "Abandon %"):
                if isinstance(val, str) and val.endswith("%"):
                    values_across_months.append(
                        float(val.replace("%", ""))
                    )
                else:
                    values_across_months.append(float(val or 0))
            elif metric_name == "Average Duration":
                avg_s = d.get("avg_duration", 0)
                values_across_months.append(avg_s)
            else:
                values_across_months.append(val)

        # Total column
        total_r = _cell_ref(ri, total_col)
        if metric_name in (
            "Total Calls", "Successful", "Failed", "Abandoned"
        ):
            total_val = sum(values_across_months)
            row_cells += (
                f'<c r="{total_r}" s="3" t="n">'
                f'<v>{total_val}</v></c>'
            )
        elif metric_name in ("Success %", "Abandon %"):
            # Store 0-1 decimal; numFmtId="10" is percentage format
            avg = (
                sum(values_across_months) / len(values_across_months)
                if values_across_months else 0
            )
            row_cells += (
                f'<c r="{total_r}" s="4" t="n">'
                f'<v>{avg:.10f}</v></c>'
            )
        elif metric_name == "Average Duration":
            avg_secs = (
                sum(values_across_months) / len(values_across_months)
                if values_across_months else 0
            )
            row_cells += (
                f'<c r="{total_r}" s="5" t="n">'
                f'<v>{round(avg_secs) / 86400:.15f}</v></c>'
            )
        else:
            row_cells += (
                f'<c r="{total_r}" s="3" t="n">'
                f'<v>{values_across_months[-1] if values_across_months else ""}</v></c>'
            )

        row_cells += "</row>"
        parts.append(row_cells)

    parts.append("</sheetData>")
    parts.append("</worksheet>")
    return "".join(parts).encode("utf-8")


def _build_monthly_sheet(month_name, paths, counts, num_paths, str_to_idx=None):
    """Build a monthly sheet (April/May/June) XML."""
    max_row = num_paths + 1  # header + data

    # Build a local path→index map if str_to_idx not provided (standalone mode)
    if str_to_idx is not None:
        path_indices = {p: str_to_idx[p] for p in paths if p in str_to_idx}
    else:
        path_indices = {p: i for i, p in enumerate(paths)}

    parts = []
    parts.append(
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<worksheet xmlns="{MAIN_NS}" '
        f'xmlns:r="{RELS_NS}">'
    )
    parts.append(f'<dimension ref="A1:B{max_row}"/>')
    parts.append('<sheetViews><sheetView workbookViewId="0"/></sheetViews>')
    parts.append('<sheetFormatPr baseColWidth="10" defaultRowHeight="15"/>')
    parts.append(
        '<cols><col min="1" max="1" width="80" customWidth="1"/>'
        '<col min="2" max="2" width="10" customWidth="1"/></cols>'
    )
    parts.append("<sheetData>")

    # Header
    parts.append(
        f'<row r="1" spans="1:2">'
        f'<c r="A1" s="1" t="s"><v>{0}</v></c>'  # "Path"
        f'<c r="B1" s="1" t="s"><v>{0}</v></c>'  # "Count"
        '</row>'
    )

    # Data rows — use actual path string indices
    for pi, (path, count) in enumerate(zip(paths, counts), start=2):
        pidx = path_indices.get(path, 0)
        parts.append(
            f'<row r="{pi}" spans="1:2">'
            f'<c r="A{pi}" s="3" t="s"><v>{pidx}</v></c>'
            f'<c r="B{pi}" t="n"><v>{count}</v></c>'
            '</row>'
        )

    parts.append("</sheetData>")
    parts.append("</worksheet>")
    return "".join(parts).encode("utf-8")


def _build_chart_xml(chart_id, month_name, paths, counts, num_paths):
    """Build chart1.xml (BarChart) with strRef for categories,
    numRef for values."""
    # Category pt elements (strRef)
    cat_pts = "".join(
        f'<c:pt idx="{i}"><c:v>{_xml_escape(paths[i])}</c:v></c:pt>'
        for i in range(num_paths)
    )

    # Value pt elements (numRef)
    val_pts = "".join(
        f'<c:pt idx="{i}"><c:v>{counts[i]}</c:v></c:pt>'
        for i in range(num_paths)
    )

    chart_height = num_paths * ROW_HEIGHT_EMU

    chart_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<ns0:chartSpace xmlns:ns0="{CHART_NS}">'
        '<ns0:chart>'
        '<ns0:viewPr/>'
        '<ns0:plotArea>'
        '<ns0:layout>'
        '<ns0:manualLayout>'
        '<ns0:anLst/>'
        '</ns0:manualLayout>'
        '</ns0:layout>'
        f'<ns0:barChart revExtraCol="0">'
        f'<ns0:grouping val="clustered"/>'
        '<ns0:ser>'
        f'<ns0:idx val="0"/>'
        f'<ns0:order val="0"/>'
        f'<ns0:xVal>'
        f'<ns0:strRef>'
        f'<ns0:strCache>'
        f'<ns0:ptCount val="{num_paths}"/>'
        f'{cat_pts}'
        f'</ns0:strCache>'
        f'</ns0:strRef>'
        f'</ns0:xVal>'
        f'<ns0:yVal>'
        f'<ns0:numRef>'
        f'<ns0:numCache>'
        '<ns0:formatCode>General</ns0:formatCode>'
        f'<ns0:ptCount val="{num_paths}"/>'
        f'{val_pts}'
        f'</ns0:numCache>'
        f'</ns0:numRef>'
        f'</ns0:yVal>'
        '</ns0:ser>'
        '<ns0:axId val="12345678"/>'
        '<ns0:axId val="87654321"/>'
        '</ns0:barChart>'
        '<ns0:catAx>'
        '<ns0:axId val="12345678"/>'
        '<ns0:scaling>'
        '<ns0:orientation val="minMax"/>'
        '</ns0:scaling>'
        '<ns0:axPos val="b"/>'
        '<ns0:majorTickMark val="none"/>'
        '<ns0:minorTickMark val="none"/>'
        '<ns0:tickLblPos val="nextTo"/>'
        '<ns0:txPr>'
        '<ns0:bodyPr rot="0" wrap="square" '
        'anchCenter="1" anchRight="1"/>'
        '<ns0:lstStyle/>'
        '<ns0:p>'
        '<ns0:pPr altLast="1" horizonTnd="0" indent="-1" rad="0"/>'
        '<ns0:r>'
        '<ns0:rPr b="0" dirty="0" fontsize="8" language="en-US"/>'
        f'<ns0:t>{_xml_escape("Call Paths")}</ns0:t>'
        '</ns0:r>'
        '</ns0:p>'
        '</ns0:txPr>'
        '</ns0:catAx>'
        '<ns0:valAx>'
        '<ns0:axId val="87654321"/>'
        '<ns0:scaling>'
        '<ns0:orientation val="minMax"/>'
        '</ns0:scaling>'
        '<ns0:axPos val="l"/>'
        '<ns0:majorGridlines/>'
        '<ns0:majorTickMark val="out"/>'
        '<ns0:minorTickMark val="none"/>'
        '<ns0:tickLblPos val="nextTo"/>'
        '<ns0:txPr>'
        '<ns0:bodyPr rot="0" wrap="square" '
        'anchCenter="1" anchRight="1"/>'
        '<ns0:lstStyle/>'
        '<ns0:p>'
        '<ns0:pPr altLast="1" horizonTnd="0" indent="-1" rad="0"/>'
        '<ns0:r>'
        '<ns0:rPr b="0" dirty="0" fontsize="8" language="en-US"/>'
        '<ns0:t>0</ns0:t>'
        '</ns0:r>'
        '</ns0:p>'
        '</ns0:txPr>'
        '</ns0:valAx>'
        '</ns0:plotArea>'
        '<ns0:legend pos="r" txtPr="barChart"/>'
        '</ns0:chart>'
        f'<ns0:chartDimensions cx="{CHART_WIDTH_EMU}" cy="{chart_height}"/>'
        '</ns0:chartSpace>'
    )
    return chart_xml.encode("utf-8")


def _build_drawing_xml(chart_file_id, month_name, num_paths, max_row=None):
    """Build drawing1.xml with BOTH oneCellAnchor (placeholder) AND
    twoCellAnchor (visible).

    Per skill spec: oneCellAnchor alone creates an invisible placeholder.
    Must include both.
    """
    OFFICE_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    chart_height = num_paths * ROW_HEIGHT_EMU
    row_to = num_paths + 1
    row_to_offset = 177_800

    drawing_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<ns0:wsDr xmlns:ns0="{DRAWING_NS}" '
        f'xmlns:r="{RELS_NS}">'
        '<ns0:nvGraphicFramePr>'
        f'<ns0:ext uri="{RELS_NS}/drawing/vnd.openxmlformats.drawingml.chartObj.xpath"/>'
        '</ns0:ext>'
        '<ns0:nvPr>'
        f'<ns0:cnvPr id="1" name="Chart {chart_file_id}" descr=""/>'
        f'<ns0:cnvGraphicFramePr/>'
        '</ns0:nvPr>'
        '</ns0:nvGraphicFramePr>'
        '<ns0:xfrm>'
        '<ns0:off x="12700" y="12700"/>'
        '<ns0:ext cx="0" cy="0"/>'
        '</ns0:xfrm>'
        '<ns0:graphicFrame>'
        '<ns0:txId>0</ns0:txId>'
        '<ns0:xfrm>'
        '<ns0:off x="495300" y="25400"/>'
        '<ns0:ext cx="0" cy="0"/>'
        '</ns0:xfrm>'
        '</ns0:graphicFrame>'
        '<ns0:clientData/>'
        # oneCellAnchor (placeholder) — per skill spec
        '<ns0:oneCellAnchor>'
        '<ns0:from x="3" y="1"/>'
        '<ns0:to x="3" y="1"/>'
        '<ns0:ext cx="0" cy="0"/>'
        '</ns0:oneCellAnchor>'
        # twoCellAnchor (visible chart) — per skill spec
        f'<ns0:twoCellAnchor>'
        f'<ns0:from x="3" y="1" xOff="12700" yOff="12700"/>'
        f'<ns0:to x="17" y="{row_to}" xOff="0" yOff="{row_to_offset}"/>'
        '<ns0:pic>'
        '<ns0:nvPicPr>'
        f'<ns0:picCF name=""/>'
        f'<ns0:picLocks noGrp="1" noMove="1" noResize="1" noSelect="1"/>'
        '<ns0:picPr>'
        f'<ns0:ext uri="{RELS_NS}/drawing/vnd.openxmlformats.drawingml.chartObj.xpath"/>'
        f'<ns0:cnvPicPr>'
        f'<ns0:cnvPr id="1" name="Chart {chart_file_id}" descr=""/>'
        f'<ns0:cnvGraphicFramePr/>'
        '</ns0:cnvPicPr>'
        '</ns0:picPr>'
        '</ns0:nvPicPr>'
        '<ns0:blipFill>'
        f'<ns0:blip r:id="rId{chart_file_id}"/>'
        '<ns0:stretch>'
        '<ns0:fillRect/>'
        '</ns0:stretch>'
        '</ns0:blipFill>'
        '<ns0:spPr>'
        '<ns0:xfrm>'
        '<ns0:off x="0" y="0"/>'
        '<ns0:ext cx="0" cy="0"/>'
        '</ns0:xfrm>'
        '<ns0:prstGeom prst="rect"/>'
        '<ns0:ln/>'
        '</ns0:spPr>'
        '<ns0:clientData/>'
        '</ns0:pic>'
        '</ns0:twoCellAnchor>'
        '</ns0:wsDr>'
    )
    return drawing_xml.encode("utf-8")


def _build_content_types(num_sheets, chart_files, drawing_files,
                         sheet_rels):
    """Build [Content_Types].xml."""
    parts = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<Types xmlns="{MAIN_NS}">'
        '<Default Extension="rels" '
        'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" '
        'ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/theme/theme1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.theme+xml"/>'
        '<Override PartName="/xl/styles.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.sharedStrings+xml"/>'
    )

    # Worksheets
    for i in range(1, num_sheets + 1):
        parts += (
            f'<Override PartName="/xl/worksheets/sheet{i}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )

    # Charts
    for i, (_, _, _) in enumerate(chart_files, start=1):
        parts += (
            f'<Override PartName="/xl/charts/chart{i}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
        )

    # Drawings
    for i, (_, _, _) in enumerate(drawing_files, start=1):
        parts += (
            f'<Override PartName="/xl/drawings/drawing{i}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.drawingml.chartshapes+xml"/>'
        )

    # Chart rels
    for i, (_, _, _) in enumerate(chart_files, start=1):
        parts += (
            f'<Override PartName="/xl/charts/_rels/chart{i}.xml.rels" '
            'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        )

    # Drawing rels
    for i, (_, _, _) in enumerate(drawing_files, start=1):
        parts += (
            f'<Override PartName="/xl/drawings/_rels/drawing{i}.xml.rels" '
            'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        )

    # Sheet rels
    for i, (_, _) in enumerate(sheet_rels, start=2):
        parts += (
            f'<Override PartName="/xl/worksheets/_rels/sheet{i}.xml.rels" '
            'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        )

    return (parts + '</Types>').encode("utf-8")


def _build_workbook(months_order):
    """Build xl/workbook.xml."""
    sheet_entries = '<sheet name="Summary" sheetId="1" r:id="rId1"/>'
    for i, m_name in enumerate(months_order):
        sheet_entries += (
            f'<sheet name="{m_name}" sheetId="{i + 2}" '
            f'r:id="rId{i + 2}"/>'
        )

    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<workbook xmlns:r="{RELS_NS}" '
        f'xmlns="{MAIN_NS}">'
        '<workbookPr/>'
        '<bookViews><workbookView showHorizontalScroll="1" '
        'showVerticalScroll="1" showSheetTabs="1" tabRatio="600" '
        'firstSheet="0" activeTab="0"/></bookViews>'
        f'<sheets>{sheet_entries}</sheets>'
        '<definedNames/><calcPr calcId="0"/></workbook>'
    ).encode("utf-8")


def _build_workbook_rels(months_order):
    """Build xl/_rels/workbook.xml.rels."""
    rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<Relationships xmlns="{RELS_NS}">'
        f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        f'<Relationship Id="rId6" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
        'Target="styles.xml"/>'
        f'<Relationship Id="rId5" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme" '
        'Target="theme/theme1.xml"/>'
        f'<Relationship Id="rId7" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
        'Target="sharedStrings.xml"/>'
    )

    for i, m_name in enumerate(months_order):
        rels += (
            f'<Relationship Id="rId{i + 2}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{i + 2}.xml"/>'
        )

    return (rels + '</Relationships>').encode("utf-8")


def _build_styles_xml():
    """Build xl/styles.xml (from reference file)."""
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<styleSheet xmlns="{MAIN_NS}">'
        '<fonts count="3">'
        '<font><sz val="11"/><name val="Calibri"/><family val="2"/></font>'
        '<font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>'
        '<font/><font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Calibri"/></font>'
        '</fonts>'
        '<fills count="3">'
        '<fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FF4472C4"/></patternFill></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FFF2F2F2"/></patternFill></fill>'
        '</fills>'
        '<borders count="1">'
        '<border><lineStyle val="thin"/></border>'
        '</borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="6">'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        '<xf numFmtId="0" fontId="1" fillId="1" borderId="0" xfId="0" applyFont="1" applyFill="1"/>'
        '<xf numFmtId="0" fontId="1" borderId="0" xfId="0" applyFont="1"/>'
        '<xf numFmtId="0" fontId="2" borderId="0" xfId="0"/>'
        '<xf numFmtId="10" fontId="2" xfId="0" applyNumberFormat="1"/>'
        '<xf numFmtId="21" fontId="2" xfId="0" applyNumberFormat="1" applyAlignment="1"><alignment horizontal="right"/></xf>'
        '</cellXfs>'
        '</styleSheet>'
    ).encode("utf-8")


def _build_theme_xml():
    """Build minimal Office theme."""
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<a:theme xmlns:a="{THEME_NS}" name="Office Theme">'
        '<a:themeElements>'
        '<a:clrScheme name="Office">'
        '<a:dk1><a:sysClr val="windowText" lastClr="000000"/></a:dk1>'
        '<a:lt1><a:sysClr val="window" lastClr="FFFFFF"/></a:lt1>'
        '<a:dk2><a:srgbClr val="1F497D"/></a:dk2>'
        '<a:lt2><a:srgbClr val="EEECE1"/></a:lt2>'
        '<a:accent1><a:srgbClr val="4472C4"/></a:accent1>'
        '<a:accent2><a:srgbClr val="E7E6E6"/></a:accent2>'
        '<a:accent3><a:srgbClr val="548235"/></a:accent3>'
        '<a:accent4><a:srgbClr val="B45D0B"/></a:accent4>'
        '<a:accent5><a:srgbClr val="BF4444"/></a:accent5>'
        '<a:accent6><a:srgbClr val="44546A"/></a:accent6>'
        '<a:hlink><a:srgbClr val="4678E6"/></a:hlink>'
        '<a:folHlink><a:srgbClr val="96607D"/></a:folHlink>'
        '</a:clrScheme>'
        '<a:fontScheme name="Calibri">'
        '<a:majorFont><a:latin typeface="Calibri"/><a:ea typeface="Calibri"/><a:cs typeface="Calibri"/></a:majorFont>'
        '<a:minorFont><a:latin typeface="Calibri"/><a:ea typeface="Calibri"/><a:cs typeface="Calibri"/></a:minorFont>'
        '</a:fontScheme>'
        '</a:themeElements>'
        '<a:objectDefaults/><a:extraClrSchemeRels/></a:theme>'
    ).encode("utf-8")


def _build_root_rels():
    """Build _rels/.rels."""
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<Relationships xmlns="{RELS_NS}">'
        f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="/xl/workbook.xml"/>'
        f'<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/metadata/core-properties" '
        'Target="/docProps/core.xml"/>'
        f'<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" '
        'Target="/docProps/app.xml"/>'
        '</Relationships>'
    ).encode("utf-8")


def _build_docprops_core():
    """Build docProps/core.xml."""
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        f'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        f'xmlns:dcterms="http://purl.org/dc/terms/" '
        f'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        f'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        '<dcterms:created xsi:type="dcterms:W3CDTF">2026-07-10T00:00:00Z</dcterms:created>'
        '<dcterms:modified xsi:type="dcterms:W3CDTF">2026-07-10T00:00:00Z</dcterms:modified>'
        '<cp:revision>1</cp:revision>'
        '</cp:coreProperties>'
    ).encode("utf-8")


def _build_docprops_app(months_order):
    """Build docProps/app.xml."""
    parts = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        f'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        '<Application>Microsoft Excel</Application>'
        '<DocSecurity>0</DocSecurity>'
        '<ScaleCrop>false</ScaleCrop>'
        '<HeadingPairs>'
        '<vt:vector size="2" baseType="variant">'
        '<vt:variant><vt:lpstr>WorkingSet</vt:variant>'
        '<vt:variant><vt:i4>' + str(len(months_order)) + '</vt:variant></vt:variant>'
        '</vt:vector>'
        '</HeadingPairs>'
        '<TitlesOfParts>'
        f'<vt:vector size="{len(months_order)}" baseType="lpstr">'
    )
    for m_name in months_order:
        parts += f'<vt:lpstr>{m_name}</vt:lpstr>'
    return (parts + '</vt:vector></TitlesOfParts></Properties>').encode("utf-8")


def _build_chart_rels(i):
    """Build chart{i}.xml.rels."""
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<Relationships xmlns="{RELS_NS}">'
        f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" '
        f'Target="../charts/chart{i}.xml"/>'
        '</Relationships>'
    ).encode("utf-8")


def _build_drawing_rels(i):
    """Build drawing{i}.xml.rels."""
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<Relationships xmlns="{RELS_NS}">'
        f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" '
        f'Target="../charts/chart{i}.xml"/>'
        '</Relationships>'
    ).encode("utf-8")
