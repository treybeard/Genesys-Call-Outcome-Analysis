#!/usr/bin/env python3
"""Process all Genesys CSV files and produce Excel workbooks."""

import csv
import os
from collections import Counter, OrderedDict
from datetime import timedelta

from openpyxl import Workbook

# Canonical project directory (scripts live in scripts/, so project root is parent)
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_DIR = os.path.join(PROJECT_ROOT, "CSV Files")
OUT_DIR = os.path.join(PROJECT_ROOT, "Output")
os.makedirs(OUT_DIR, exist_ok=True)

# Batch definitions - using actual cache document names
BATCHES = {
    "Switchboard": [
        "doc_b696083cdb96_Switchbaord-April.csv",
        "doc_91b7d70ec942_Switchbaord-May.csv",
        "doc_73661aaf8493_Switchboard June.csv",
    ],
    "Non-ATTY": [
        "doc_d98d6ef6429d_Non-ATTY April.csv",
        "doc_65dfd2d8d7f7_Non-ATTY May.csv",
        "doc_14b5e25c2c25_Non-ATTY June.csv",
    ],
    "ATTY": [
        "doc_d0cec11475c8_ATTY April.csv",
        "doc_38dfa13d8eb9_ATTY May.csv",
        "doc_74e69e181d95_ATTY June.csv",
    ],
    "Bar Examiners": [
        "doc_c848269b044c_Bar Examiners April.csv",
        "doc_7ec1019bfcbe_Bar Examiners May.csv",
        "doc_a7d02a33e3f6_Bar Examiners June.csv",
    ],
    "LRS Spanish": [
        "doc_87985000e90b_LRS Spanish April.csv",
        "doc_66820355f921_LRS Spanish May.csv",
        "doc_edceaaca07cc_LRS Spanish June.csv",
    ],
}


def parse_duration(raw):
    """Parse duration string like ' 00:04:12.710' into seconds. Returns None if invalid or >24h."""
    if not raw or not isinstance(raw, str):
        return None
    raw = raw.strip()
    if not raw:
        return None
    parts = raw.split(":")
    if len(parts) == 3:
        try:
            h, m, s = parts[0].strip(), parts[1].strip(), parts[2].strip()
            total = int(h) * 3600 + int(m) * 60 + float(s)
            if total > 86400:  # > 24 hours = data error
                return None
            return total
        except (ValueError, IndexError):
            return None
    return None


def analyze_csv(filepath, is_lrs_spanish=False):
    """Analyze a single Genesys CSV file."""
    total_calls = 0
    successful = 0
    failed = 0
    abandoned = 0
    durations = []
    path_counter = Counter()

    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            total_calls += 1

            # Outcome classification
            outcome_success = row.get("Outcome Success", "").strip() == "1"
            abandoned_flag = row.get("Abandoned", "").strip() == "YES"

            if is_lrs_spanish:
                if outcome_success and not abandoned_flag:
                    successful += 1
                elif abandoned_flag:
                    abandoned += 1
                else:
                    failed += 1
            else:
                if outcome_success:
                    successful += 1
                elif abandoned_flag:
                    abandoned += 1
                else:
                    failed += 1

            # Duration
            dur_raw = row.get("Duration", "")
            dur_secs = parse_duration(dur_raw)
            if dur_secs is not None:
                durations.append(dur_secs)

            # Queue paths
            queue_raw = row.get("Queue", "")
            if queue_raw and str(queue_raw).strip():
                paths = [p.strip() for p in str(queue_raw).split(";") if p.strip()]
                if paths:
                    path_key = " → ".join(paths)
                    path_counter[path_key] += 1

    avg_duration = sum(durations) / len(durations) if durations else 0

    return {
        "total_calls": total_calls,
        "successful": successful,
        "failed": failed,
        "abandoned": abandoned,
        "success_pct": (successful / total_calls * 100) if total_calls else 0,
        "abandon_pct": (abandoned / total_calls * 100) if total_calls else 0,
        "avg_duration": avg_duration,
        "path_counter": path_counter,
    }


def write_excel(filepath, name, metrics, batch_name):
    """Write Excel workbook with Summary and Call Paths sheets."""
    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Metric"
    ws["B1"] = "Value"

    avg_dur_str = str(timedelta(seconds=round(metrics["avg_duration"]))) if metrics["avg_duration"] > 0 else "N/A"

    summary_rows = [
        ("Call Path", batch_name),
        ("Month", name),
        ("Total Calls", metrics["total_calls"]),
        ("Successful", metrics["successful"]),
        ("Failed", metrics["failed"]),
        ("Abandoned", metrics["abandoned"]),
        ("Success %", f"{metrics['success_pct']:.2f}%"),
        ("Abandon %", f"{metrics['abandon_pct']:.2f}%"),
        ("Average Duration", avg_dur_str),
    ]

    for i, (metric, value) in enumerate(summary_rows, start=2):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)

    # Call Paths sheet
    ws2 = wb.create_sheet("Call Paths")
    ws2["A1"] = "Path"
    ws2["B1"] = "Count"

    sorted_paths = sorted(metrics["path_counter"].items(), key=lambda x: x[1], reverse=True)

    for i, (path, count) in enumerate(sorted_paths, start=2):
        ws2.cell(row=i, column=1, value=path)
        ws2.cell(row=i, column=2, value=count)

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws2.column_dimensions["A"].width = 60
    ws2.column_dimensions["B"].width = 10

    wb.save(filepath)


# Main processing loop
for batch_name, filenames in BATCHES.items():
    print(f"\n{'='*60}")
    print(f"Processing batch: {batch_name}")
    print(f"{'='*60}")

    is_lrs = batch_name == "LRS Spanish"

    for fname in filenames:
        csv_path = os.path.join(CSV_DIR, fname)
        if not os.path.exists(csv_path):
            print(f"  ⚠ Missing: {fname}")
            continue

        print(f"\n  Processing: {fname}")
        metrics = analyze_csv(csv_path, is_lrs_spanish=is_lrs)

        # Create friendly output name from the doc_ prefixed filename
        clean_name = fname
        # Strip the doc_ prefix and UUID
        if clean_name.startswith("doc_"):
            parts = clean_name.split("_", 1)
            if len(parts) == 2:
                # Extract the original filename after the doc_ prefix
                # e.g., "doc_b696083cdb96_Switchbaord-April.csv" -> "Switchbaord-April.csv"
                after_uuid = parts[1]
                # The format is doc_<uuid>_<original_filename>
                # We need to find where the UUID ends (32 hex chars)
                hex_part = after_uuid.split("_")[0] if "_" in after_uuid else ""
                if len(hex_part) == 32 and all(c in "0123456789abcdef" for c in hex_part):
                    clean_name = after_uuid[len(hex_part)+1:]  # skip hex + underscore
        
        clean_name = clean_name.replace(".csv", "")
        # Handle the typo in the filename
        clean_name = clean_name.replace("Switchbaord", "Switchboard")

        out_name = f"{batch_name} - {clean_name}.xlsx"
        out_path = os.path.join(OUT_DIR, out_name)

        write_excel(out_path, clean_name, metrics, batch_name)

        print(f"    Total: {metrics['total_calls']:,}")
        print(f"    Success: {metrics['successful']:,} ({metrics['success_pct']:.1f}%)")
        print(f"    Failed: {metrics['failed']:,}")
        print(f"    Abandoned: {metrics['abandoned']:,} ({metrics['abandon_pct']:.1f}%)")
        print(f"    Avg Duration: {timedelta(seconds=round(metrics['avg_duration']))}")
        print(f"    Unique Paths: {len(metrics['path_counter'])}")
        print(f"    ✓ Saved: {out_name}")

print(f"\n{'='*60}")
print(f"Done! All files saved to: {OUT_DIR}")
print(f"{'='*60}")
