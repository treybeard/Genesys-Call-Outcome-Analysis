#!/usr/bin/env python3
"""
Genesys Call Flow Report Consolidator
======================================
Combines monthly individual Call Flow report files into consolidated
April–June summary reports with formatted data and bar charts.

Usage:
    python3 consolidate_call_flow_reports.py

Requires:
    pandas, openpyxl (already in Python environment)

Input:
    Individual monthly .xlsx files in the project directory, e.g.:
    - Switchboard April.xlsx, Switchboard May.xlsx, Switchboard June.xlsx
    - Non-ATTY April.xlsx, Non-ATTY May.xlsx, Non-ATTY June.xlsx
    - ATTY April.xlsx, ATTY May.xlsx, ATTY June.xlsx
    - Bar Examiners April.xlsx, Bar Examiners May.xlsx, Bar Examiners June.xlsx
    - LRS Spanish April.xlsx, LRS Spanish May.xlsx, LRS Spanish June.xlsx

Output:
    Consolidated April-June .xlsx files with:
    - Summary sheet (April | May | June | Total columns)
    - Individual monthly sheets (April, May, June) with bar charts
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─── Configuration ───────────────────────────────────────────────────────

PROJECTS_DIR = '/Users/trey/.hermes/projects/Genesys Call Analysis/'

CATEGORIES = {
    'Switchboard': ['Switchboard April.xlsx', 'Switchboard May.xlsx', 'Switchboard June.xlsx'],
    'Non-ATTY': ['Non-ATTY April.xlsx', 'Non-ATTY May.xlsx', 'Non-ATTY June.xlsx'],
    'ATTY': ['ATTY April.xlsx', 'ATTY May.xlsx', 'ATTY June.xlsx'],
    'Bar Examiners': ['Bar Examiners April.xlsx', 'Bar Examiners May.xlsx', 'Bar Examiners June.xlsx'],
    'LRS Spanish': ['LRS Spanish April.xlsx', 'LRS Spanish May.xlsx', 'LRS Spanish June.xlsx'],
}

MONTH_ORDER = ['April', 'May', 'June']

OUTPUT_SUFFIX = 'April-June.xlsx'


# ─── Styles ──────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


# ─── Consolidation Logic ────────────────────────────────────────────────

def load_month_data(filenames):
    """Load Summary and Call Paths sheets from each month's file."""
    all_summary = []
    all_paths = []
    labels = []
    
    for i, (filename, label) in enumerate(zip(filenames, MONTH_ORDER)):
        filepath = os.path.join(PROJECTS_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  ⚠ Missing: {filename}")
            continue
        
        summary_df = pd.read_excel(filepath, sheet_name='Summary')
        call_paths_df = pd.read_excel(filepath, sheet_name='Call Paths')
        
        all_summary.append((summary_df, label))
        all_paths.append((call_paths_df, label))
        labels.append(label)
        print(f"  ✓ Loaded {filename}")
    
    return all_summary, all_paths


def build_consolidated_summary(all_summary):
    """Build combined Summary with Monthly columns + Total."""
    metrics = [
        'Total Calls', 'Successful', 'Success %', 'Failed', 'Failure %',
        'Abandoned', 'Abandon %', 'Average Call Time (MM:SS)',
    ]
    
    combined = {'Metric': metrics}
    
    # Monthly columns
    for label in MONTH_ORDER:
        values = []
        for metric in metrics:
            found = False
            for df, lbl in all_summary:
                if lbl == label and metric in df['Metric'].values:
                    val = df.loc[df['Metric'] == metric, 'Value'].iloc[0]
                    values.append(val)
                    found = True
                    break
            values.append(values[-1] if found else 'N/A')
        combined[label] = values
    
    # Total column (sum for counts, weighted avg for %)
    total_t = 0
    success_t = 0
    failed_t = 0
    abandon_t = 0
    weighted_dur_sum = 0.0
    weighted_dur_count = 0
    
    for df, _ in all_summary:
        calls = df.loc[df['Metric'] == 'Total Calls', 'Value'].iloc[0]
        succ = df.loc[df['Metric'] == 'Successful', 'Value'].iloc[0]
        fail = df.loc[df['Metric'] == 'Failed', 'Value'].iloc[0]
        aband = df.loc[df['Metric'] == 'Abandoned', 'Value'].iloc[0]
        
        # Need duration in seconds for weighted calc
        duration_seconds = df.loc[
            df['Metric'] == 'Average Duration (seconds)', 'Value'
        ].iloc[0] if 'Average Duration (seconds)' in df['Metric'].values else 0
        
        total_t += calls
        success_t += succ
        failed_t += fail
        abandon_t += aband
        weighted_dur_sum += duration_seconds * calls
        weighted_dur_count += calls
    
    total_values = [
        total_t,                                          # Total Calls
        success_t,                                        # Successful
        round(success_t / total_t * 100, 2) if total_t else 0,  # Success %
        failed_t,                                         # Failed
        round(failed_t / total_t * 100, 2) if total_t else 0,   # Failure %
        abandon_t,                                        # Abandoned
        round(abandon_t / total_t * 100, 2) if total_t else 0,  # Abandon %
        # Average Call Time (MM:SS)
        _format_duration(weighted_dur_sum / weighted_dur_count) if weighted_dur_count else "N/A",
    ]
    
    combined['Total'] = total_values
    return pd.DataFrame(combined)


def _format_duration(total_seconds):
    """Convert seconds to mm:ss format."""
    minutes = int(total_seconds // 60)
    seconds = round(total_seconds % 60)
    if seconds >= 60:
        minutes += 1
        seconds = 0
    return f"{minutes}:{seconds:02d}"


def build_consolidated_paths(all_paths):
    """Concatenate all months' Call Paths and group+sum."""
    consolidated = pd.concat([df for df, _ in all_paths], ignore_index=True)
    consolidated = consolidated.groupby('Queue Path', as_index=False)['Call Count'].sum()
    consolidated = consolidated.sort_values('Call Count', ascending=False).reset_index(drop=True)
    return consolidated


# ─── Excel Formatting ───────────────────────────────────────────────────

def format_summary_sheet(ws):
    """Apply formatting to the Summary sheet."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
            if col == ws.max_column and row > 1:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
    
    ws.column_dimensions['A'].width = 35
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def format_monthly_sheet(ws):
    """Apply formatting to a monthly Call Path sheet."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15


# ─── Chart Creation ─────────────────────────────────────────────────────

def add_monthly_chart(ws, sheet_name):
    """Add a column bar chart to a monthly Call Path sheet."""
    chart = BarChart()
    chart.title = f"{sheet_name} Call Paths"
    chart.y_axis.title = "Call Count"
    chart.x_axis.title = "Queue Path"
    chart.style = 10
    chart.width = 15
    chart.height = 12
    
    # Data starts from row 2 (skip header row)
    labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    
    anchor_row = ws.max_row + 5
    ws.add_chart(chart, f"A{anchor_row}")


# ─── Main Pipeline ──────────────────────────────────────────────────────

def process_category(cat_name, filenames, output_path):
    """Run full consolidation pipeline for one category."""
    print(f"\n{'='*60}")
    print(f"Processing: {cat_name}")
    print(f"{'='*60}")
    
    # Load all months
    all_summary, all_paths = load_month_data(filenames)
    if not all_summary:
        print("  ❌ No data to consolidate, skipping.")
        return False
    
    # Build consolidated Summary
    consolidated_summary = build_consolidated_summary(all_summary)
    print(f"  Summary: {len(consolidated_summary)} metrics × 4 columns")
    
    # Build consolidated individual monthly sheets (for separate monthly views)
    monthly_data = {}
    for df, label in all_paths:
        monthly_data[label] = df
    
    # Save workbook
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        consolidated_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Save each month as individual sheet (these will get charts)
        for label, df in monthly_data.items():
            df.to_excel(writer, sheet_name=label, index=False)
    
    # Now apply formatting and add charts via openpyxl
    wb = load_workbook(output_path)
    
    # Format Summary
    if 'Summary' in wb.sheetnames:
        format_summary_sheet(wb['Summary'])
        print(f"  ✓ Formatted Summary sheet")
    
    # Format monthly sheets + add charts
    for label in MONTH_ORDER:
        if label in wb.sheetnames:
            format_monthly_sheet(wb[label])
            add_monthly_chart(wb[label], label)
            print(f"  ✓ Formatted + charted {label} sheet ({len(wb[label]) - 1} paths)")
    
    # Save final
    wb.save(output_path)
    print(f"  ✓ Saved: {output_path}")
    return True


def main():
    """Run consolidation pipeline for all categories."""
    print("=" * 60)
    print("Genesys Call Flow Report Consolidator")
    print(f"Project directory: {PROJECTS_DIR}")
    print("=" * 60)
    
    results = {}
    
    for cat_name, filenames in CATEGORIES.items():
        output_filename = f"{cat_name} {OUTPUT_SUFFIX}".replace('-', ' ').strip()
        output_path = os.path.join(PROJECTS_DIR, output_filename)
        
        try:
            success = process_category(cat_name, filenames, output_path)
            results[cat_name] = '✓' if success else '✗'
        except Exception as e:
            print(f"  ❌ Error processing {cat_name}: {e}")
            import traceback
            traceback.print_exc()
            results[cat_name] = '✗'
    
    # Summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    for cat, status in results.items():
        print(f"  {status} {cat}")
    
    success_count = sum(1 for s in results.values() if s == '✓')
    print(f"\n✅ {success_count}/{len(results)} categories processed successfully")


if __name__ == '__main__':
    main()
